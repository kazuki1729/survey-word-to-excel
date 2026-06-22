"""
アンケート調査票（Word .docx）を、集計・分析しやすい Excel（.xlsx）へ変換するツール。

使い方:
    python convert.py 入力.docx 出力.xlsx

引数を省略した場合は sample/ のサンプルを変換します:
    python convert.py
"""

import re
import sys
from pathlib import Path

from docx import Document
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# 全角数字 → 半角数字
_Z2H = str.maketrans("０１２３４５６７８９", "0123456789")

# 設問の行：「問1.」「Q1）」「Ｑ1：」など
QUESTION_RE = re.compile(r"^\s*(?:問|Q|Ｑ)\s*([0-9０-９]+)\s*[\.\．、）\):：]?\s*(.*)$")
# 選択肢の行：「1.」「2）」など番号＋区切り
OPTION_NUM_RE = re.compile(r"^\s*([0-9０-９]+)\s*[\.\．、）\)]\s*(.+)$")
# 選択肢の行：丸数字「①」など
OPTION_CIRCLE_RE = re.compile(r"^\s*([①-⑳])\s*(.+)$")
# 選択肢の行：「ア．」「イ）」などカタカナ
OPTION_KANA_RE = re.compile(r"^\s*([ア-ン])\s*[\.\．、）\)]\s*(.+)$")


def classify(line):
    """1行を (種類, 内容...) に分類する。"""
    q = QUESTION_RE.match(line)
    if q:
        num = q.group(1).translate(_Z2H)
        return ("question", num, q.group(2).strip())
    for rx in (OPTION_NUM_RE, OPTION_KANA_RE):
        m = rx.match(line)
        if m:
            return ("option", m.group(2).strip())
    m = OPTION_CIRCLE_RE.match(line)
    if m:
        return ("option", m.group(2).strip())
    return ("other", line.strip())


def parse_docx(path):
    """Word を読み、[{番号, 設問文, 選択肢[...]}] のリストにする。"""
    doc = Document(path)
    questions = []
    current = None
    for para in doc.paragraphs:
        line = para.text.strip()
        if not line:
            continue
        kind, *rest = classify(line)
        if kind == "question":
            num, text = rest
            current = {"番号": num, "設問文": text, "選択肢": []}
            questions.append(current)
        elif kind == "option" and current is not None:
            current["選択肢"].append(rest[0])
        elif kind == "other" and current is not None:
            # 設問文が空なら次の行を設問文として補う。
            if not current["設問文"]:
                current["設問文"] = rest[0]
            else:
                # 補足（「複数選択可」等）は設問文に追記。
                current["設問文"] += " " + rest[0]
    return questions


def write_xlsx(questions, path):
    """ワイド形式とロング形式の2シートで Excel を出力する。"""
    header_fill = PatternFill("solid", fgColor="1D9E75")
    header_font = Font(bold=True, color="FFFFFF")
    wrap = Alignment(wrap_text=True, vertical="top")

    wb = Workbook()

    # --- シート1：ワイド形式（1設問1行＋回答欄） ---
    ws = wb.active
    ws.title = "調査票（ワイド形式）"
    max_opts = max((len(q["選択肢"]) for q in questions), default=0)
    headers = ["設問番号", "設問文"] + [f"選択肢{i}" for i in range(1, max_opts + 1)] + ["回答"]
    ws.append(headers)
    for q in questions:
        row = [q["番号"], q["設問文"]] + q["選択肢"] + [""] * (max_opts - len(q["選択肢"])) + [""]
        ws.append(row)
    _style_sheet(ws, headers, header_fill, header_font, wrap, text_col=2)

    # --- シート2：ロング形式（分析用・1選択肢1行） ---
    ws2 = wb.create_sheet("選択肢一覧（ロング形式）")
    headers2 = ["設問番号", "設問文", "選択肢番号", "選択肢内容"]
    ws2.append(headers2)
    for q in questions:
        if q["選択肢"]:
            for i, opt in enumerate(q["選択肢"], start=1):
                ws2.append([q["番号"], q["設問文"], i, opt])
        else:
            ws2.append([q["番号"], q["設問文"], "", "（自由記述）"])
    _style_sheet(ws2, headers2, header_fill, header_font, wrap, text_col=2)

    wb.save(path)


def _style_sheet(ws, headers, fill, font, wrap, text_col):
    """ヘッダー装飾・列幅・折り返し・ウィンドウ枠の固定。"""
    for c, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        if col == text_col:
            ws.column_dimensions[letter].width = 50
        elif col == 1:
            ws.column_dimensions[letter].width = 10
        else:
            ws.column_dimensions[letter].width = 20
    for row in ws.iter_rows(min_row=2):
        row[text_col - 1].alignment = wrap
    ws.freeze_panes = "A2"


def main():
    if len(sys.argv) >= 3:
        src, dst = Path(sys.argv[1]), Path(sys.argv[2])
    else:
        src = Path("sample/sample_survey.docx")
        dst = Path("sample/survey.xlsx")
        print("引数が無いため、サンプルを変換します。")

    if not src.exists():
        sys.exit(f"入力ファイルが見つかりません: {src}")

    questions = parse_docx(src)
    write_xlsx(questions, dst)
    total_opts = sum(len(q["選択肢"]) for q in questions)
    print(f"変換完了: {src} → {dst}")
    print(f"  設問数: {len(questions)} 問 / 選択肢: 計 {total_opts} 件")


if __name__ == "__main__":
    main()
